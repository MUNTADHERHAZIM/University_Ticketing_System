"""
نظام التقارير المتقدم والمراقبة المباشرة
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q, Count, Avg, F, Sum
from datetime import timedelta, datetime
from .models import Ticket, TicketAction, TicketAcknowledgment
from accounts.models import CustomUser, PenaltyPoints, Department
from .decorators import can_view_reports, can_view_monitoring
import json
import csv


@login_required
@can_view_monitoring
def monitoring_dashboard(request):
    """
    لوحة المراقبة المباشرة - Real-time Dashboard
    """
    now = timezone.now()
    
    # الطلبات المتأخرة بالوقت الفعلي
    overdue_tickets = Ticket.objects.filter(
        status__in=['new', 'pending_ack', 'in_progress'],
        sla_deadline__lt=now
    ).select_related('department', 'assigned_to', 'created_by').order_by('sla_deadline')
    
    # الطلبات الحرجة
    critical_tickets = Ticket.objects.filter(
        priority='critical',
        status__in=['new', 'pending_ack', 'in_progress']
    ).select_related('department', 'assigned_to').order_by('sla_deadline')
    
    # مؤشرات الأداء الرئيسية (KPIs)
    total_tickets = Ticket.objects.count()
    pending_tickets = Ticket.objects.filter(status__in=['new', 'pending_ack', 'in_progress']).count()
    violated_tickets = Ticket.objects.filter(status='violated').count()
    resolved_today = Ticket.objects.filter(
        resolved_at__gte=now.replace(hour=0, minute=0, second=0)
    ).count()
    
    # نسبة الالتزام بالمهلة
    total_resolved = Ticket.objects.filter(status__in=['resolved', 'closed']).count()
    on_time_resolved = Ticket.objects.filter(
        status__in=['resolved', 'closed'],
        resolved_at__lte=F('sla_deadline')
    ).count()
    compliance_rate = (on_time_resolved / total_resolved * 100) if total_resolved > 0 else 0
    
    # متوسط وقت الاستجابة (بالساعات)
    avg_response = Ticket.objects.filter(
        acknowledged_at__isnull=False
    ).annotate(
        response_time=F('acknowledged_at') - F('created_at')
    ).aggregate(avg=Avg('response_time'))
    
    avg_response_hours = 0
    if avg_response['avg']:
        avg_response_hours = avg_response['avg'].total_seconds() / 3600
    
    # متوسط وقت الحل (بالساعات)
    avg_resolution = Ticket.objects.filter(
        resolved_at__isnull=False
    ).annotate(
        resolution_time=F('resolved_at') - F('created_at')
    ).aggregate(avg=Avg('resolution_time'))
    
    avg_resolution_hours = 0
    if avg_resolution['avg']:
        avg_resolution_hours = avg_resolution['avg'].total_seconds() / 3600
    
    # الأقسام الأسوأ أداءً
    worst_departments = Department.objects.annotate(
        violated_count=Count('tickets', filter=Q(tickets__status='violated')),
        pending_count=Count('tickets', filter=Q(tickets__status__in=['new', 'pending_ack', 'in_progress']))
    ).order_by('-violated_count')[:5]
    
    context = {
        'overdue_tickets': overdue_tickets,
        'critical_tickets': critical_tickets,
        'total_tickets': total_tickets,
        'pending_tickets': pending_tickets,
        'violated_tickets': violated_tickets,
        'resolved_today': resolved_today,
        'compliance_rate': round(compliance_rate, 1),
        'avg_response_hours': round(avg_response_hours, 1),
        'avg_resolution_hours': round(avg_resolution_hours, 1),
        'worst_departments': worst_departments,
    }
    
    return render(request, 'tickets/monitoring_dashboard.html', context)


@login_required
@can_view_monitoring
def monitoring_api(request):
    """
    API للحصول على بيانات المراقبة المباشرة (للتحديث التلقائي)
    """
    now = timezone.now()
    
    # عدد الطلبات المتأخرة
    overdue_count = Ticket.objects.filter(
        status__in=['new', 'pending_ack', 'in_progress'],
        sla_deadline__lt=now
    ).count()
    
    # عدد الطلبات الحرجة
    critical_count = Ticket.objects.filter(
        priority='critical',
        status__in=['new', 'pending_ack', 'in_progress']
    ).count()
    
    # عدد الطلبات المعلقة
    pending_count = Ticket.objects.filter(
        status__in=['new', 'pending_ack', 'in_progress']
    ).count()
    
    return JsonResponse({
        'overdue_count': overdue_count,
        'critical_count': critical_count,
        'pending_count': pending_count,
        'timestamp': now.isoformat()
    })


@login_required
@can_view_reports
def performance_report(request):
    """
    تقرير الأداء الشامل
    """
    now = timezone.now()
    
    # فترة التقرير (افتراضياً آخر 30 يوم)
    period_days = int(request.GET.get('period', 30))
    start_date = now - timedelta(days=period_days)
    
    # إحصائيات عامة
    total_tickets = Ticket.objects.filter(created_at__gte=start_date).count()
    resolved_tickets = Ticket.objects.filter(
        resolved_at__gte=start_date,
        status__in=['resolved', 'closed']
    ).count()
    violated_tickets = Ticket.objects.filter(
        created_at__gte=start_date,
        status='violated'
    ).count()
    
    # أداء الأقسام
    departments_performance = Department.objects.annotate(
        total=Count('tickets', filter=Q(tickets__created_at__gte=start_date)),
        resolved=Count('tickets', filter=Q(
            tickets__resolved_at__gte=start_date,
            tickets__status__in=['resolved', 'closed']
        )),
        violated=Count('tickets', filter=Q(
            tickets__created_at__gte=start_date,
            tickets__status='violated'
        )),
        penalty_points=Sum('penalties__points', filter=Q(penalties__created_at__gte=start_date))
    ).order_by('-violated')
    
    # أداء الموظفين
    employees_performance = CustomUser.objects.filter(
        role__in=['employee', 'head']
    ).annotate(
        total=Count('assigned_tickets', filter=Q(assigned_tickets__created_at__gte=start_date)),
        resolved=Count('assigned_tickets', filter=Q(
            assigned_tickets__resolved_at__gte=start_date,
            assigned_tickets__status__in=['resolved', 'closed']
        )),
        violated=Count('assigned_tickets', filter=Q(
            assigned_tickets__created_at__gte=start_date,
            assigned_tickets__status='violated'
        )),
        penalty_points=Sum('penalties__points', filter=Q(penalties__created_at__gte=start_date))
    ).filter(total__gt=0).order_by('-violated')
    
    # أفضل الموظفين
    best_employees = employees_performance.filter(violated=0).order_by('-resolved')[:10]
    
    # أسوأ الموظفين
    worst_employees = employees_performance.filter(violated__gt=0).order_by('-violated')[:10]
    
    context = {
        'period_days': period_days,
        'start_date': start_date,
        'total_tickets': total_tickets,
        'resolved_tickets': resolved_tickets,
        'violated_tickets': violated_tickets,
        'resolution_rate': round((resolved_tickets / total_tickets * 100) if total_tickets > 0 else 0, 1),
        'violation_rate': round((violated_tickets / total_tickets * 100) if total_tickets > 0 else 0, 1),
        'departments_performance': departments_performance,
        'best_employees': best_employees,
        'worst_employees': worst_employees,
    }
    
    return render(request, 'tickets/performance_report.html', context)


@login_required
@can_view_reports
def export_performance_excel(request):
    """
    تصدير تقرير الأداء إلى Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('مكتبة openpyxl غير مثبتة', status=500)
    
    now = timezone.now()
    period_days = int(request.GET.get('period', 30))
    start_date = now - timedelta(days=period_days)
    
    # إنشاء workbook
    wb = openpyxl.Workbook()
    
    # ورقة الأقسام
    ws_dept = wb.active
    ws_dept.title = "أداء الأقسام"
    
    # العناوين
    headers = ['القسم', 'إجمالي الطلبات', 'المحلولة', 'المخالفة', 'النقاط الجزائية']
    ws_dept.append(headers)
    
    # تنسيق العناوين
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws_dept[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # البيانات
    departments = Department.objects.annotate(
        total=Count('tickets', filter=Q(tickets__created_at__gte=start_date)),
        resolved=Count('tickets', filter=Q(
            tickets__resolved_at__gte=start_date,
            tickets__status__in=['resolved', 'closed']
        )),
        violated=Count('tickets', filter=Q(
            tickets__created_at__gte=start_date,
            tickets__status='violated'
        )),
        penalty_points=Sum('penalties__points', filter=Q(penalties__created_at__gte=start_date))
    ).order_by('-violated')
    
    for dept in departments:
        ws_dept.append([
            dept.name,
            dept.total,
            dept.resolved,
            dept.violated,
            dept.penalty_points or 0
        ])
    
    # ورقة الموظفين
    ws_emp = wb.create_sheet("أداء الموظفين")
    headers = ['الموظف', 'القسم', 'إجمالي الطلبات', 'المحلولة', 'المخالفة', 'النقاط الجزائية']
    ws_emp.append(headers)
    
    for cell in ws_emp[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    employees = CustomUser.objects.filter(
        role__in=['employee', 'head']
    ).annotate(
        total=Count('assigned_tickets', filter=Q(assigned_tickets__created_at__gte=start_date)),
        resolved=Count('assigned_tickets', filter=Q(
            assigned_tickets__resolved_at__gte=start_date,
            assigned_tickets__status__in=['resolved', 'closed']
        )),
        violated=Count('assigned_tickets', filter=Q(
            assigned_tickets__created_at__gte=start_date,
            assigned_tickets__status='violated'
        )),
        penalty_points=Sum('penalties__points', filter=Q(penalties__created_at__gte=start_date))
    ).filter(total__gt=0).order_by('-violated')
    
    for emp in employees:
        ws_emp.append([
            emp.get_full_name(),
            emp.department.name if emp.department else 'غير محدد',
            emp.total,
            emp.resolved,
            emp.violated,
            emp.penalty_points or 0
        ])
    
    # ضبط عرض الأعمدة
    for ws in [ws_dept, ws_emp]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # حفظ الملف
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="performance_report_{now.strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@can_view_reports
def penalty_points_report(request):
    """
    تقرير النقاط الجزائية
    """
    now = timezone.now()
    period_days = int(request.GET.get('period', 30))
    start_date = now - timedelta(days=period_days)
    
    # النقاط الجزائية حسب الموظف
    employee_penalties = CustomUser.objects.annotate(
        total_points=Sum('penalties__points', filter=Q(penalties__created_at__gte=start_date)),
        penalty_count=Count('penalties', filter=Q(penalties__created_at__gte=start_date))
    ).filter(total_points__gt=0).order_by('-total_points')
    
    # النقاط الجزائية حسب القسم
    department_penalties = Department.objects.annotate(
        total_points=Sum('penalties__points', filter=Q(penalties__created_at__gte=start_date)),
        penalty_count=Count('penalties', filter=Q(penalties__created_at__gte=start_date))
    ).filter(total_points__gt=0).order_by('-total_points')
    
    # تصنيف الموظفين
    excellent = employee_penalties.filter(total_points__lte=5)
    good = employee_penalties.filter(total_points__gt=5, total_points__lte=15)
    acceptable = employee_penalties.filter(total_points__gt=15, total_points__lte=30)
    poor = employee_penalties.filter(total_points__gt=30, total_points__lte=50)
    very_poor = employee_penalties.filter(total_points__gt=50)
    
    context = {
        'period_days': period_days,
        'employee_penalties': employee_penalties,
        'department_penalties': department_penalties,
        'excellent': excellent,
        'good': good,
        'acceptable': acceptable,
        'poor': poor,
        'very_poor': very_poor,
    }
    
    return render(request, 'tickets/penalty_points_report.html', context)
